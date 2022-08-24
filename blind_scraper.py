from bs4 import BeautifulSoup
import requests
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
import time
from threading import Timer 
from datetime import datetime
from interruptingcow import timeout



def body_text(url):
    text =[]
    try:
        check = str(url)
        index = check.find(".pdf")
        if index == -1:
            html_text = requests.get(url).text
            soup = BeautifulSoup(html_text, 'lxml')

            article_texts = soup.find_all('p')
            for sent in article_texts:
                text.append(sent.getText())
    except:
        text.append("request has been blocked!")

    return str(text)


result = pd.read_excel (r"read the links that you savd by runing main.py")

wb = load_workbook(r"where you would like to have a file to add the result of blind scraping")
sheet = wb.active

sheet.cell(row=1, column=5).value="text from blind scraper"


def functionALI (i):
    url = result.iloc[i,3]

    now = datetime.now()
    start_time = time.time()
    current_time = now.strftime("%H:%M:%S")
    
    sheet.cell(row= i+2 , column=5).value = body_text(url)
    wb.save("US1.xlsx")
    
    return print("link \t" + str(i) + " \t Nicely Scraped *___* "), print("start runing =", current_time), start_time



from func_timeout import func_timeout, FunctionTimedOut
for j in tqdm(range(0,3700)):
    try:
        doitReturnValue = func_timeout(5, functionALI, args=(j,))

    except FunctionTimedOut:
        print ( str(j) + "  could not complete within 5 seconds and was terminated.\n" )


